"""
Script to generate Excel task tracker from big-data.md plan
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Define members and their tasks per week
members_data = {
    "Member 1 - Leader": {
        "role": "Team Leader + Infrastructure + Java Backend",
        "skills": "Java, Python, Kafka, Kubernetes, System Integration",
        "weeks": {
            1: {
                "goal": "Infrastructure Foundation",
                "tasks": [
                    ("Setup Kafka cluster", "Confluent Cloud hoặc Local với Docker. Tạo topics: orders_raw, orders_enriched, revenue_metrics. Test producer/consumer"),
                    ("Setup MongoDB Atlas", "Tạo database olist_analytics. Tạo collections: revenue_metrics, customer_segments, product_performance"),
                    ("Setup MinIO", "S3-compatible storage. Tạo buckets: bronze, silver, gold. Test upload/download"),
                    ("Setup GitHub repository", "Tạo folder structure theo README. Setup branch strategy (dev, feature branches)"),
                    ("Document connection strings", "Document connection strings và credentials"),
                ],
                "deliverables": "Kafka topics ready; MongoDB collections created; MinIO buckets ready; GitHub repo structure; config/ folder với connection configs"
            },
            2: {
                "goal": "Kubernetes & Integration",
                "tasks": [
                    ("Setup Minikube", "Install Minikube. Start cluster. Verify kubectl working"),
                    ("Create Kubernetes manifests", "deployment/k8s/namespace.yaml, debezium-deployment.yaml, configmap.yaml"),
                    ("Setup Airflow", "docker-compose for Airflow. Access Airflow UI. Create first DAG"),
                    ("Integration testing", "Test Kafka → MongoDB connectivity. Test MinIO read/write from Spark"),
                    ("Code review setup", "Setup PR template. Define code review process"),
                ],
                "deliverables": "Minikube running; K8s manifests created; Airflow operational; Integration tests passing"
            },
            3: {
                "goal": "Monitoring & Code Review",
                "tasks": [
                    ("Setup monitoring framework", "Install Prometheus (optional). Configure Spark metrics. Create metrics dashboard template"),
                    ("Code review for all members", "Review Spark code quality. Check naming conventions. Verify error handling"),
                    ("Create integration test framework", "tests/integration/ structure. Pytest setup. Sample test cases"),
                    ("Weekly progress meeting", "Review deliverables. Identify blockers. Adjust timeline if needed"),
                ],
                "deliverables": "Monitoring framework setup; Code reviews completed; Integration test framework ready"
            },
            4: {
                "goal": "Performance Monitoring",
                "tasks": [
                    ("Setup Spark UI monitoring", "Access Spark UI (port 4040). Analyze job stages. Identify shuffle operations"),
                    ("Resource allocation tuning", "Tune Spark configs. Document optimal settings"),
                    ("Cost analysis", "Estimate cloud costs. Resource usage report. Optimization recommendations"),
                    ("Code review optimization work", "Review Member 3's optimization work"),
                ],
                "deliverables": "Spark UI monitoring documented; Optimal Spark configs identified; Cost analysis report"
            },
            5: {
                "goal": "Airflow DAGs for Streaming",
                "tasks": [
                    ("Create Airflow DAG for stream monitoring", "Check Kafka consumer lag. Check Spark streaming job status. Alert if lag > threshold"),
                    ("Integration testing", "End-to-end test: Postgres → Kafka → Spark → MongoDB. Data consistency checks. Latency measurements"),
                    ("Documentation", "Update architecture diagram. Document data flows. API documentation"),
                ],
                "deliverables": "Airflow monitoring DAG; Integration tests passing; Documentation updated"
            },
            6: {
                "goal": "Kubernetes Deployment",
                "tasks": [
                    ("Deploy Spark Streaming on K8s", "Create Spark on K8s YAML. Deploy streaming job. Test auto-restart on failure"),
                    ("Setup CI/CD pipeline", "GitHub Actions for build. Automated tests. Docker image build"),
                    ("Load testing streaming pipeline", "Generate high event rate. Monitor Kafka lag. Measure end-to-end latency"),
                    ("Document deployment process", "Deployment guide. Troubleshooting tips"),
                ],
                "deliverables": "Streaming job on Kubernetes; CI/CD pipeline (basic); Load testing results; Deployment documentation"
            },
            7: {
                "goal": "Service Integration",
                "tasks": [
                    ("Implement service discovery", "Kubernetes Services. Service mesh (optional). DNS-based discovery"),
                    ("Implement load balancing", "Kafka partition assignment. Consumer group management. Load balancer config"),
                    ("End-to-end integration testing", "Test full pipeline. Data consistency validation. Performance testing. Failure recovery testing"),
                    ("Create integration test suite", "Automated E2E tests. Test data generator. Validation scripts"),
                ],
                "deliverables": "Service discovery working; Load balancing configured; E2E integration tests passing; Integration test suite created"
            },
            8: {
                "goal": "MLflow Setup & Code Review",
                "tasks": [
                    ("Setup MLflow", "Install MLflow. Configure tracking server. Model registry"),
                    ("Code review for ML code", "Review model training code. Verify evaluation metrics. Check for overfitting"),
                    ("Create model deployment guide", "How to load models. How to make predictions. Model versioning strategy"),
                    ("Integration testing with ML models", "Test batch prediction. Test real-time prediction"),
                ],
                "deliverables": "MLflow setup (optional); ML code review completed; Model deployment guide; ML integration tests"
            },
            9: {
                "goal": "Integration & Performance Testing",
                "tasks": [
                    ("End-to-end integration testing", "Data flow validation. Data consistency checks. Latency measurements. Throughput testing"),
                    ("Performance tuning", "Identify bottlenecks. Optimize slow queries. Tune Spark configurations. Optimize Kafka settings"),
                    ("Load testing", "Generate high data volume. Stress test all components. Measure breaking points. Document results"),
                    ("Create deployment runbook", "Step-by-step deployment guide. Configuration checklist. Troubleshooting guide. Rollback procedures"),
                ],
                "deliverables": "E2E integration tests passing; Performance tuning completed; Load testing report; Deployment runbook"
            },
            10: {
                "goal": "Final Report & Presentation",
                "tasks": [
                    ("Write final report (30-40 pages)", "Introduction, Architecture, Tech Stack, Implementation, Spark Features, Testing, Monitoring, Challenges, Future Work"),
                    ("Create presentation slides (20-25)", "Title, Problem, Architecture, Spark Features DEMO, Streaming, ML, Results, Challenges, Demo, Future, Q&A"),
                    ("Record demo video (10-15 min)", "Introduction. Architecture walkthrough. Live demo. Grafana dashboards. Code walkthrough. Q&A prep"),
                    ("Prepare Q&A document", "Anticipate questions. Prepare answers. Technical deep dives ready"),
                ],
                "deliverables": "Final report (30-40 pages); Presentation slides (20-25 slides); Demo video (10-15 min); Q&A document"
            },
        }
    },
    "Member 2 - Java Dev": {
        "role": "Java Backend Engineer + CDC Specialist",
        "skills": "Java, Spring Boot, Kafka, PostgreSQL, Debezium",
        "weeks": {
            1: {
                "goal": "PostgreSQL & Development Environment",
                "tasks": [
                    ("Setup PostgreSQL database", "Install PostgreSQL 15+. Enable WAL for CDC. Create database olist_db"),
                    ("Create Java Spring Boot project", "Initialize Maven/Gradle project. Add dependencies: Spring Boot, Kafka, MongoDB, PostgreSQL. Setup application.properties template"),
                    ("Study Debezium CDC concepts", "Read Debezium documentation. Prepare CDC configuration template"),
                    ("Install development tools", "Java 17, Maven, IntelliJ IDEA. Kafka command-line tools"),
                ],
                "deliverables": "PostgreSQL installed và configured for CDC; Java Spring Boot project initialized; Development environment ready"
            },
            2: {
                "goal": "Data Loading & CDC Setup",
                "tasks": [
                    ("Load Olist data vào PostgreSQL", "Create tables schema (9 tables). Write data loading scripts. Load augmented dataset (2-3GB). Verify data integrity"),
                    ("Setup Debezium CDC", "Configure Debezium connector for PostgreSQL. Deploy Debezium trên Kubernetes. Test CDC capturing changes"),
                    ("Create Java Kafka Producer (test)", "Simple producer để test Kafka. Send test messages. Verify messages in Kafka topic"),
                ],
                "deliverables": "PostgreSQL tables created and loaded with data; Debezium CDC capturing changes; Java Kafka producer working"
            },
            3: {
                "goal": "Assist Batch + Start Java Consumer",
                "tasks": [
                    ("Assist Member 3 with Spark SQL", "Review join strategies. Help debug performance issues"),
                    ("Start Java Kafka Consumer project", "services/java-consumer/ project. Spring Boot + Kafka consumer. Deserialize messages. Log messages"),
                    ("Study MongoDB Java driver", "Add MongoDB dependency. Create connection class. Test CRUD operations"),
                    ("Create error handling framework", "Dead letter queue (DLQ) pattern. Retry logic. Circuit breaker (Resilience4j)"),
                ],
                "deliverables": "Java Kafka Consumer reading messages; MongoDB Java driver working; Error handling framework implemented"
            },
            4: {
                "goal": "MongoDB Writer Service",
                "tasks": [
                    ("Create MongoDB writer microservice", "services/java-mongodb-writer/ Spring Boot. REST API. Kafka consumer. Write to MongoDB"),
                    ("Implement batch write optimization", "Bulk insert API. Connection pooling. Retry mechanism"),
                    ("Add health check endpoint", "/health endpoint. Check Kafka connection. Check MongoDB connection"),
                    ("Dockerize Java services", "Create Dockerfile. docker-compose for local testing. Push to Docker Hub (optional)"),
                ],
                "deliverables": "Java MongoDB writer service working; Kafka → Java → MongoDB pipeline operational; Docker images created"
            },
            5: {
                "goal": "Kafka Producer Enhancement",
                "tasks": [
                    ("Enhance Java Kafka Producer", "Read from PostgreSQL periodically. Produce to Kafka topics. Transactional writes (exactly-once)"),
                    ("Implement CDC listener", "PostgreSQL logical replication. Capture INSERT/UPDATE/DELETE. Produce events to Kafka"),
                    ("Add metrics", "Prometheus metrics export. Messages produced per second. Error count"),
                    ("Load testing", "Generate high volume of events. Measure throughput. Identify bottlenecks"),
                ],
                "deliverables": "Kafka producer producing events; Metrics exported; Load testing results"
            },
            6: {
                "goal": "Circuit Breaker & Resilience",
                "tasks": [
                    ("Implement Circuit Breaker pattern", "Use Resilience4j. Circuit breaker for MongoDB writes. Circuit breaker for Kafka producer"),
                    ("Implement retry logic", "Exponential backoff. Max retry attempts. Dead letter queue (DLQ)"),
                    ("Add comprehensive logging", "Structured logging (JSON). Log levels configuration. Correlation IDs for tracing"),
                    ("Integration testing", "Test failure scenarios. Verify circuit breaker opens. Verify retry logic works"),
                ],
                "deliverables": "Circuit breaker implemented; Retry logic working; Comprehensive logging added; Integration tests passing"
            },
            7: {
                "goal": "Java Consumer & Producer Complete",
                "tasks": [
                    ("Complete Java Kafka Consumer", "Consume from multiple topics. Parallel processing. Offset management. Consumer group config"),
                    ("Complete Java MongoDB Writer", "Batch writes. Bulk upsert. Connection pooling. Write performance optimization"),
                    ("Implement data enrichment service", "Read from Kafka. Enrich with reference data. Write enriched data to Kafka. Add product details to orders"),
                    ("Add monitoring endpoints", "/metrics for Prometheus. /health endpoint. /info endpoint. Consumer lag metrics"),
                    ("Performance tuning", "JVM tuning. Thread pool sizing. Memory management. Benchmark results"),
                ],
                "deliverables": "Java Kafka consumer operational; Java MongoDB writer operational; Data enrichment service working; Monitoring endpoints available; Performance tuning completed"
            },
            8: {
                "goal": "ML Model Serving (Optional)",
                "tasks": [
                    ("Create model serving API", "REST API to serve predictions. Load ML model (PMML/ONNX). Prediction endpoint"),
                    ("Assist Member 5 with ML", "Help with data pipeline for ML. Review ML code"),
                    ("Finalize Java services documentation", "API documentation. Deployment guide. Configuration guide"),
                    ("Performance testing Java services", "Load testing. Stress testing. Benchmark results"),
                ],
                "deliverables": "ML serving API (optional); ML infrastructure support provided; Java documentation completed; Performance testing done"
            },
            9: {
                "goal": "Monitoring & Logging",
                "tasks": [
                    ("Centralized logging setup", "Configure structured logging. Log aggregation. Log correlation IDs"),
                    ("Metrics collection", "Expose Prometheus metrics from Java. JVM metrics. Application metrics"),
                    ("Health checks", "Kubernetes liveness probes. Readiness probes. Dependency health checks"),
                    ("Testing", "Integration tests for Java services. Error scenario testing. Performance regression tests"),
                ],
                "deliverables": "Centralized logging working; Metrics collection operational; Health checks configured; Testing completed"
            },
            10: {
                "goal": "Java Services Documentation",
                "tasks": [
                    ("Write Java services documentation", "Architecture of Java services. API documentation (OpenAPI/Swagger). Configuration guide. Deployment guide"),
                    ("Code documentation", "Javadoc for all classes. README for each service. Example usage"),
                    ("Create deployment guide", "Docker deployment. Kubernetes deployment. Configuration management"),
                    ("Prepare demo scenarios", "Java Kafka consumer demo. MongoDB writer performance demo. Circuit breaker demo"),
                ],
                "deliverables": "Java documentation completed; API documentation; Deployment guide; Demo scenarios prepared"
            },
        }
    },
    "Member 3 - Spark Batch": {
        "role": "PySpark Batch Processing Specialist",
        "skills": "PySpark, Data Transformation, Performance Optimization",
        "weeks": {
            1: {
                "goal": "Dataset & Spark Setup",
                "tasks": [
                    ("Download full Olist dataset", "Download từ Kaggle. Verify data integrity. Place in data/external/. All 9 CSV files"),
                    ("Setup PySpark environment", "Install Java 17, Python 3.12, Spark 3.5.8. Configure SPARK_HOME, JAVA_HOME. Test Spark local mode"),
                    ("Data exploration & profiling", "Load all 9 CSV files. Check schema, nulls, data types. Create data dictionary. Write exploration notebook"),
                    ("Create data augmentation script", "Script to duplicate data 10x. Add timestamp variations. Target: 2-3GB dataset"),
                ],
                "deliverables": "All 9 CSV files downloaded; PySpark working locally; Data exploration report; Data augmentation script"
            },
            2: {
                "goal": "Bronze Layer Ingestion",
                "tasks": [
                    ("Create data_ingestion.py", "Read 9 CSV files. Validate schemas. Handle data quality issues. Write to Bronze layer (MinIO) as Parquet. Partition by date"),
                    ("Create data quality framework", "Check nulls, duplicates. Schema validation. Data profiling stats. Generate quality report"),
                    ("Create Airflow DAG", "DAG: batch_ingestion_dag.py. Schedule: daily. Error handling"),
                    ("Test incremental loading", "Support append mode. Deduplication logic"),
                ],
                "deliverables": "Bronze layer populated with Parquet files; Data quality report; Airflow DAG working; Incremental loading tested"
            },
            3: {
                "goal": "Joins & UDFs ⭐",
                "tasks": [
                    ("Implement Broadcast Join", "orders ⋈ product_categories (small table). Use broadcast() function"),
                    ("Implement Sort-Merge Join", "orders ⋈ order_items ⋈ products (large tables). Repartition and sortWithinPartitions"),
                    ("Implement Multiple Join", "orders ⋈ items ⋈ products ⋈ sellers ⋈ customers. Analyze execution plans"),
                    ("Create 3+ custom UDFs", "Calculate discount percentage. Classify delivery status. Customer segment classifier"),
                    ("Write Silver layer", "Partition by year and month. Parquet with Snappy compression. Schema evolution handling"),
                ],
                "deliverables": "Broadcast join implemented and working; Sort-merge join implemented; 3+ UDFs created and tested; Silver layer data written; Execution plan analysis document"
            },
            4: {
                "goal": "Advanced Transformations ⭐",
                "tasks": [
                    ("Implement Window Functions", "Ranking: Top 10 products per category. Lead/Lag: Month-over-month growth. Running Totals: Cumulative spend"),
                    ("Implement Pivot/Unpivot", "Pivot: Payment methods by state. Unpivot: using stack function"),
                    ("Implement Partitioning & Bucketing", "Partition by year, month. Bucket frequently joined tables"),
                    ("Implement Caching Strategies", "Cache frequently accessed DataFrames. Persist with storage level. Unpersist when done"),
                    ("Performance benchmarking", "Benchmark before optimization. Apply optimizations. Benchmark after. Document improvements"),
                ],
                "deliverables": "Window functions implemented (3+ examples); Pivot/Unpivot working; Partitioned data in MinIO; Bucketed tables created; Caching strategies applied; Performance improvement report (before/after metrics)"
            },
            5: {
                "goal": "Support Streaming",
                "tasks": [
                    ("Assist Member 4 with Streaming", "Debug streaming queries. Help with transformations"),
                    ("Optimize batch jobs", "Apply learnings from week 4. Refactor code for reusability. Add unit tests"),
                    ("Create reusable Spark utilities", "utils/spark_utils.py. Common transformations. UDF library. Configuration loader"),
                ],
                "deliverables": "Streaming support provided; Batch jobs optimized; Utility library created"
            },
            6: {
                "goal": "Testing & Documentation",
                "tasks": [
                    ("Write unit tests", "Test UDFs. Test transformations. Test aggregations. Use pytest + pyspark test utils"),
                    ("Write integration tests", "End-to-end batch pipeline test. Data quality tests. Performance regression tests"),
                    ("Code refactoring", "Extract common functions. Improve code readability. Add docstrings"),
                    ("Documentation", "Batch processing guide. Performance tuning guide. API documentation"),
                ],
                "deliverables": "Unit tests (>80% coverage); Integration tests passing; Code refactored; Documentation completed"
            },
            7: {
                "goal": "Batch-Stream Integration",
                "tasks": [
                    ("Create unified data model", "Reconcile batch and streaming schemas. Merge batch and streaming results in MongoDB. Deduplication logic"),
                    ("Implement Lambda serving layer", "Batch view + Real-time view merge. Query API (Flask/FastAPI). Serve queries from MongoDB"),
                    ("Create data reconciliation job", "Compare batch and streaming results. Identify discrepancies. Generate reconciliation report"),
                    ("Performance comparison", "Batch vs streaming latency. Resource utilization. Cost analysis"),
                ],
                "deliverables": "Unified data model; Lambda serving layer; Reconciliation job working; Performance comparison report"
            },
            8: {
                "goal": "Support ML Data Prep",
                "tasks": [
                    ("Assist Member 5 with data prep", "Feature engineering pipeline. Data sampling for training. Data validation"),
                    ("Optimize batch jobs for ML", "Efficient feature extraction. Caching intermediate results"),
                    ("Create data export utilities", "Export data for ML (CSV, Parquet). Sampling utilities. Data versioning"),
                    ("Code cleanup and refactoring", "Remove dead code. Improve modularity. Add more unit tests"),
                ],
                "deliverables": "ML data prep support completed; Batch jobs optimized for ML; Data export utilities created; Code cleaned up"
            },
            9: {
                "goal": "Testing & Documentation",
                "tasks": [
                    ("Comprehensive testing", "Unit tests for all functions. Integration tests for pipelines. Data quality tests. Performance regression tests"),
                    ("Test coverage report", "Measure code coverage (>80% target). Identify untested code. Add missing tests"),
                    ("Documentation", "Batch processing user guide. Code documentation (docstrings). Performance tuning guide. Troubleshooting guide"),
                    ("Create demo scenarios", "Demo 1: Batch processing pipeline. Demo 2: Performance optimization. Demo 3: Data quality checks"),
                ],
                "deliverables": "Test coverage >80%; All tests passing; Documentation completed; Demo scenarios prepared"
            },
            10: {
                "goal": "Batch Processing Documentation",
                "tasks": [
                    ("Write batch documentation", "Architecture of batch pipeline. Data flow explanation. Performance optimization guide. Troubleshooting guide"),
                    ("Create code documentation", "Docstrings for all functions. README for each module. Example usage"),
                    ("Create demo notebook", "Jupyter notebook demonstrating: Broadcast join, Window functions, Pivot operations, Performance optimization"),
                    ("Contribute to final report", "Batch processing section. Performance metrics section"),
                ],
                "deliverables": "Batch documentation completed; Code documentation; Demo notebook; Report contribution"
            },
        }
    },
    "Member 4 - Spark Streaming": {
        "role": "PySpark Streaming Specialist",
        "skills": "Spark Structured Streaming, Real-time Processing, Watermarking",
        "weeks": {
            1: {
                "goal": "Streaming Study & Prep",
                "tasks": [
                    ("Study Spark Structured Streaming", "Read official documentation. Complete streaming tutorials. Understand window operations"),
                    ("Study Kafka integration with Spark", "Kafka-Spark connector. Schema management. Offset management"),
                    ("Prepare streaming test data generator", "Python script to produce fake orders to Kafka. Configurable rate (orders/second)"),
                    ("Install development tools", "Python 3.12, PySpark 3.5.8. Kafka Python client"),
                ],
                "deliverables": "Streaming concepts documented; Test data generator script ready; Development environment ready"
            },
            2: {
                "goal": "Kafka Consumer Setup",
                "tasks": [
                    ("Create kafka_consumer.py", "Read from Kafka topic orders_raw. Parse JSON messages. Schema definition. Write to console (testing)"),
                    ("Test streaming pipeline basics", "Use test data generator. Verify Spark can read from Kafka. Check offset management"),
                    ("Create checkpoint directory", "Setup checkpoint location in MinIO. Test recovery after failure"),
                    ("Document streaming architecture", "Data flow diagram. Kafka topics mapping"),
                ],
                "deliverables": "Kafka consumer reading messages; Checkpoint mechanism working; Streaming architecture document"
            },
            3: {
                "goal": "Assist Batch + Streaming Prep",
                "tasks": [
                    ("Assist Member 3 with batch", "Help debug join issues. Performance testing"),
                    ("Prepare streaming transformations", "Design window aggregations. Plan watermarking strategy. Schema for streaming data"),
                    ("Create streaming data generator", "Produce realistic order events to Kafka. Configurable event rate. Late arrival simulation (10% delayed)"),
                    ("Test Kafka-Spark streaming", "Read from Kafka in streaming mode. Write to console. Verify latency"),
                ],
                "deliverables": "Batch processing support completed; Streaming data generator ready; Streaming connection tested"
            },
            4: {
                "goal": "Streaming Architecture",
                "tasks": [
                    ("Design streaming architecture doc", "Data flow diagram. Topic naming conventions. Schema registry strategy"),
                    ("Prepare watermarking test scenarios", "Late data simulation. Different watermark thresholds. Impact analysis"),
                    ("Create monitoring dashboard template", "Kafka lag metrics. Processing rate. Error rate"),
                    ("Study stateful streaming", "mapGroupsWithState API. flatMapGroupsWithState API. State timeout mechanisms"),
                ],
                "deliverables": "Streaming architecture document; Test scenarios prepared; Monitoring dashboard template; Stateful streaming concepts documented"
            },
            5: {
                "goal": "Basic Streaming ⭐",
                "tasks": [
                    ("Read from Kafka and parse", "Create streaming_aggregations.py. Read from orders_raw. Parse JSON messages"),
                    ("Implement Tumbling Window", "Revenue per 5 minutes. With watermark (10 minutes)"),
                    ("Implement Sliding Window", "Moving 1-hour revenue (slide 10 minutes). With watermark"),
                    ("Write to MongoDB (streaming)", "Use foreachBatch. Output mode: update"),
                    ("Write to MinIO", "Checkpoint + data. Trigger: processingTime 1 minute"),
                    ("Test output modes", "Append, Update, Complete modes"),
                ],
                "deliverables": "Kafka → Spark Streaming working; Tumbling window implemented; Sliding window implemented; Streaming → MongoDB working; Streaming → MinIO working; Output modes tested"
            },
            6: {
                "goal": "Advanced Streaming ⭐",
                "tasks": [
                    ("Implement Watermarking", "Allow 10 minutes of late data. Test with late data injection. Measure data loss. Document strategy"),
                    ("Implement Session Windows", "Session window: gap of 30 minutes. Customer session analysis"),
                    ("Implement Stateful Streaming", "mapGroupsWithState for session tracking. Update user session state"),
                    ("Implement Exactly-Once", "Kafka transactional writes. Idempotent writes to MongoDB. Test duplicate detection"),
                    ("Checkpoint management", "Optimize checkpoint interval. Test recovery. Monitor checkpoint size"),
                ],
                "deliverables": "Watermarking implemented and tested; Session windows working; Stateful streaming implemented; Exactly-once semantics verified; Checkpoint recovery tested"
            },
            7: {
                "goal": "Streaming Optimization",
                "tasks": [
                    ("Optimize streaming performance", "Tune Kafka consumer settings. Optimize trigger interval. Minimize shuffle operations. Coalesce/Repartition"),
                    ("Implement streaming analytics", "Real-time fraud detection (simple rules). Anomaly detection (statistical). Trending products detection"),
                    ("Add stream monitoring", "Processing rate metrics. Batch processing time. Watermark delay metrics. State size monitoring"),
                    ("Failure recovery testing", "Kill streaming job. Verify checkpoint recovery. Measure recovery time. Document recovery procedure"),
                ],
                "deliverables": "Streaming performance optimized; Real-time analytics implemented; Monitoring metrics added; Failure recovery tested and documented"
            },
            8: {
                "goal": "Real-time Feature Engineering",
                "tasks": [
                    ("Real-time feature computation", "Streaming aggregations as features. Rolling window features. Stateful features"),
                    ("Write features to feature store", "Real-time feature updates to MongoDB. Feature versioning"),
                    ("Create feature serving API", "FastAPI to query features. Low-latency retrieval. Caching layer"),
                    ("Assist Member 5 with ML", "Help with streaming ML inference. Review ML code"),
                ],
                "deliverables": "Real-time feature engineering; Features in feature store; Feature serving API; ML support provided"
            },
            9: {
                "goal": "Streaming Testing",
                "tasks": [
                    ("Streaming integration tests", "Test window aggregations. Test watermarking with late data. Test stateful processing. Test exactly-once"),
                    ("Failure recovery testing", "Test checkpoint recovery. Test after Kafka restart. Test after Spark restart. Measure recovery time"),
                    ("Performance testing", "Measure processing latency. Measure throughput. Test with different data rates. Document results"),
                    ("Documentation", "Streaming user guide. Troubleshooting guide. Demo scenarios"),
                ],
                "deliverables": "Streaming tests passing; Failure recovery verified; Performance testing completed; Documentation done"
            },
            10: {
                "goal": "Streaming Documentation",
                "tasks": [
                    ("Write streaming documentation", "Architecture of streaming pipeline. Watermarking strategy. Stateful processing guide. Troubleshooting guide"),
                    ("Create code documentation", "Docstrings for all functions. README for streaming module. Example usage"),
                    ("Create demo notebook", "Jupyter notebook: Window aggregations, Watermarking, Stateful processing, Exactly-once semantics"),
                    ("Contribute to final report", "Streaming processing section. Real-time metrics section"),
                ],
                "deliverables": "Streaming documentation completed; Code documentation; Demo notebook; Report contribution"
            },
        }
    },
    "Member 5 - Data Scientist": {
        "role": "ML Engineer + Data Analyst + Visualization Engineer",
        "skills": "MLlib, GraphFrames, Grafana, MongoDB",
        "weeks": {
            1: {
                "goal": "ML Study & MongoDB Setup",
                "tasks": [
                    ("Study Spark MLlib", "ALS for recommendation. Random Forest for classification. Regression models"),
                    ("Study GraphFrames", "PageRank algorithm. Connected Components. Graph construction"),
                    ("Setup MongoDB connection from Python", "Install pymongo. Test CRUD operations. Create MongoDB schema design document"),
                    ("Install Grafana locally", "Install Grafana. Connect to MongoDB. Create test dashboard"),
                ],
                "deliverables": "ML study notes; MongoDB connection working from Python; MongoDB schema design document; Grafana installed and connected"
            },
            2: {
                "goal": "MongoDB Integration",
                "tasks": [
                    ("Create mongo_connector.py", "Connection manager class. CRUD operations wrappers. Bulk insert optimizations. Upsert logic"),
                    ("Create MongoDB indexes", "Indexes on collections. Compound indexes. TTL indexes for hot/cold data"),
                    ("Write data quality checker", "scripts/data_quality.py. Null checks. Duplicate detection. Outlier detection. Generate HTML report"),
                    ("Test MongoDB write performance", "Benchmark bulk inserts. Query performance testing. Document results"),
                ],
                "deliverables": "MongoDB connector working; Indexes created; Data quality checker ready; Performance benchmark report"
            },
            3: {
                "goal": "Gold Layer Aggregations",
                "tasks": [
                    ("Create silver_to_gold.py", "Revenue aggregations by category, year_month. Sum, count, avg, countDistinct"),
                    ("Implement RFM Segmentation", "Recency: days since last purchase. Frequency: number of orders. Monetary: total spend"),
                    ("Product Performance Metrics", "Top products by revenue. Products by review score. Conversion rate by product"),
                    ("Write Gold data to MongoDB", "Use mongo-spark connector. Upsert mode. Batch write optimization"),
                    ("Create data quality dashboard", "Grafana dashboard. Null percentages. Record counts. Data freshness"),
                ],
                "deliverables": "Gold layer aggregations created; Data written to MongoDB; Data quality dashboard in Grafana"
            },
            4: {
                "goal": "Feature Engineering for ML",
                "tasks": [
                    ("Prepare ML datasets", "Recommendation dataset (user-item matrix). Churn prediction dataset (RFM features). Revenue forecasting dataset (time series features)"),
                    ("Feature engineering pipeline", "services/ml/feature_engineering.py. Scalers and encoders. Handle missing values. Train/test split"),
                    ("Exploratory Data Analysis", "Correlation analysis. Distribution plots. Feature importance (preliminary). Create EDA notebook"),
                    ("Write features to MongoDB", "Collection: ml_features. Versioning strategy"),
                ],
                "deliverables": "ML datasets prepared; Feature engineering pipeline; EDA notebook; Features stored in MongoDB"
            },
            5: {
                "goal": "MongoDB Query Optimization",
                "tasks": [
                    ("Analyze query patterns", "Slow query log analysis. Identify frequent queries. Measure query latency"),
                    ("Create additional indexes", "Compound indexes for common queries. Text indexes for search. Geospatial indexes (if needed)"),
                    ("Implement aggregation pipelines", "MongoDB aggregation for dashboards. Pre-compute heavy aggregations. Materialized views pattern"),
                    ("Start Grafana dashboards", "Dashboard 1: Business Metrics (draft). Connect MongoDB. Create basic panels"),
                ],
                "deliverables": "Query optimization completed; Indexes created; Aggregation pipelines implemented; First Grafana dashboard (draft)"
            },
            6: {
                "goal": "Grafana Dashboards Development",
                "tasks": [
                    ("Create Dashboard 2: Customer Analytics", "Customer segments pie chart. RFM heatmap. Top customers table. CLV trend"),
                    ("Create Dashboard 3: Product Performance", "Top products bar chart. Product revenue trend. Review score distribution. Category breakdown"),
                    ("Add alerting rules", "Alert if revenue drops > 20%. Alert if error rate > 5%. Alert if Kafka lag > 1000"),
                    ("Dashboard optimization", "Query optimization. Caching strategies. Auto-refresh configuration"),
                ],
                "deliverables": "Dashboard 2 completed; Dashboard 3 completed; Alerting rules configured; Dashboards optimized"
            },
            7: {
                "goal": "Prepare for ML",
                "tasks": [
                    ("Finalize ML datasets", "Feature selection. Train/validation/test split. Data normalization/scaling"),
                    ("Setup ML pipeline structure", "services/ml/ folder. Model training scripts. Model evaluation framework"),
                    ("Create feature store in MongoDB", "Collection: feature_store. Versioning mechanism. Feature metadata"),
                    ("Dashboard 4: System Monitoring", "Kafka metrics. Spark metrics. MongoDB metrics. JVM metrics. System metrics"),
                ],
                "deliverables": "ML datasets finalized; ML pipeline structure ready; Feature store in MongoDB; Dashboard 4: System Monitoring completed"
            },
            8: {
                "goal": "ML & Advanced Analytics ⭐",
                "tasks": [
                    ("Train ALS Recommendation model", "User-item matrix. Train ALS. Evaluate RMSE. Generate top-N recommendations"),
                    ("Train Churn Prediction model", "Feature engineering. Train Random Forest. Evaluate AUC, Precision, Recall. Feature importance"),
                    ("Train Revenue Forecasting model", "Time series features. Linear Regression or GBT. Evaluate RMSE, R²"),
                    ("GraphFrames analysis", "Construct seller-product-customer graph. PageRank: influential sellers. Connected Components. Triangle Count. Motif finding"),
                    ("Time series analysis", "Seasonality decomposition. Trend analysis. Anomaly detection (IQR method)"),
                ],
                "deliverables": "3 ML models trained and evaluated; Model performance report; GraphFrames analysis (PageRank, Components, Triangles); Time series analysis; Models saved to MinIO/MongoDB; Jupyter notebooks"
            },
            9: {
                "goal": "Finalize Dashboards & Visualization",
                "tasks": [
                    ("Finalize all Grafana dashboards", "Dashboard 1-4 polish. Dashboard 5: ML Model Performance (new)"),
                    ("Add interactive features", "Drill-down capabilities. Date range selectors. Variable templates"),
                    ("Configure alerts", "Revenue anomaly alerts. System health alerts. Data quality alerts. Email/Slack notifications"),
                    ("Create dashboard user guide", "How to navigate. What each metric means. How to interpret alerts"),
                    ("ML model monitoring dashboard", "Model performance over time. Prediction accuracy. Feature drift detection"),
                ],
                "deliverables": "5 Grafana dashboards finalized; Interactive features added; Alerts configured; Dashboard user guide; ML monitoring dashboard"
            },
            10: {
                "goal": "ML & Visualization Documentation",
                "tasks": [
                    ("Write ML documentation", "Model architecture & algorithms. Feature engineering explained. Model evaluation & results. GraphFrames analysis. Time series analysis"),
                    ("Create ML notebooks", "Notebook 1: ALS Recommendation. Notebook 2: Churn Prediction. Notebook 3: Revenue Forecasting. Notebook 4: Graph Analytics. Notebook 5: Time Series"),
                    ("Write dashboard documentation", "Dashboard user guide. Metrics explanation. How to create alerts"),
                    ("Contribute to final report", "ML & Analytics section. Visualization section. Results & insights section"),
                ],
                "deliverables": "ML documentation completed; 5 ML notebooks; Dashboard documentation; Report contribution"
            },
        }
    },
}

def create_excel_plan():
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    week_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    week_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create data validation for status column
    status_validation = DataValidation(type="list", formula1='"PENDING,IMPLEMENT,DONE"', allow_blank=False)
    status_validation.error = 'Your entry is not in the list'
    status_validation.errorTitle = 'Invalid Entry'
    
    # Create sheet for each member
    for member_name, member_info in members_data.items():
        # Create sheet
        ws = wb.create_sheet(title=member_name[:31])  # Excel sheet name max 31 chars
        
        # Add member info header
        ws.merge_cells('A1:E1')
        ws['A1'] = f"{member_name}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Vai trò: {member_info['role']}"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:E3')
        ws['A3'] = f"Kỹ năng: {member_info['skills']}"
        ws['A3'].font = Font(italic=True, size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Add column headers
        row_num = 5
        headers = ['Tuần', 'Công việc', 'Status', 'Mô tả', 'Deliverables']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 8   # Tuần
        ws.column_dimensions['B'].width = 50  # Công việc
        ws.column_dimensions['C'].width = 12  # Status
        ws.column_dimensions['D'].width = 80  # Mô tả
        ws.column_dimensions['E'].width = 80  # Deliverables
        
        row_num += 1
        
        # Add data for each week
        for week_num in range(1, 11):
            if week_num in member_info['weeks']:
                week_data = member_info['weeks'][week_num]
                
                # Week header row
                ws.merge_cells(f'A{row_num}:E{row_num}')
                week_cell = ws.cell(row=row_num, column=1, value=f"TUẦN {week_num}: {week_data['goal']}")
                week_cell.fill = week_fill
                week_cell.font = week_font
                week_cell.alignment = Alignment(horizontal='left', vertical='center')
                week_cell.border = border
                ws.row_dimensions[row_num].height = 20
                row_num += 1
                
                # Add tasks
                for task_name, task_desc in week_data['tasks']:
                    # Week number
                    week_cell = ws.cell(row=row_num, column=1, value=week_num)
                    week_cell.alignment = Alignment(horizontal='center', vertical='top')
                    week_cell.border = border
                    
                    # Task name
                    task_cell = ws.cell(row=row_num, column=2, value=task_name)
                    task_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    task_cell.border = border
                    
                    # Status (with dropdown)
                    status_cell = ws.cell(row=row_num, column=3, value="PENDING")
                    status_cell.alignment = Alignment(horizontal='center', vertical='top')
                    status_cell.border = border
                    status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    status_validation.add(status_cell)
                    
                    # Description
                    desc_cell = ws.cell(row=row_num, column=4, value=task_desc)
                    desc_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    desc_cell.border = border
                    
                    # Deliverables (same for all tasks in the week)
                    deliv_cell = ws.cell(row=row_num, column=5, value=week_data['deliverables'])
                    deliv_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    deliv_cell.border = border
                    
                    ws.row_dimensions[row_num].height = 60
                    row_num += 1
        
        # Add data validation
        ws.add_data_validation(status_validation)
        
        # Freeze panes
        ws.freeze_panes = 'A6'
    
    # Save workbook
    output_file = r"C:\Work\Big-Data\plan.xlsx"
    wb.save(output_file)
    print(f"Excel file created successfully: {output_file}")

if __name__ == "__main__":
    create_excel_plan()
